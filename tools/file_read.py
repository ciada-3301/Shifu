from langchain_core.tools import tool
from pathlib import Path

PLAYGROUND_DIR = Path("")
PLAYGROUND_DIR.mkdir(exist_ok=True)
@tool
def file_read(filepath: str) -> str:
    """
    Read any supported file and return its text content.
    Supported: plain text, source code, markdown, JSON, CSV, YAML,
               PDF (.pdf — needs pypdf),
               Word (.docx — needs python-docx),
               Excel (.xlsx/.xls — needs openpyxl).
    Paths that are NOT absolute are resolved relative to Playground/.
    """
    path = Path(filepath)
    if not path.is_absolute():
        path = PLAYGROUND_DIR / path
    if not path.exists():
        return f"Error: file not found — {path}"
 
    ext = path.suffix.lower()
 
    if ext == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            pages  = [pg.extract_text() or "" for pg in reader.pages]
            return "\n\n".join(pages).strip() or "(PDF: no extractable text)"
        except ImportError:
            return "Error: PDF support requires  pip install pypdf"
        except Exception as e:
            return f"Error reading PDF: {e}"
 
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(str(path))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip()) or "(empty)"
        except ImportError:
            return "Error: Word support requires  pip install python-docx"
        except Exception as e:
            return f"Error reading DOCX: {e}"
 
    if ext in (".xlsx", ".xls", ".xlsm"):
        try:
            import openpyxl
            wb    = openpyxl.load_workbook(str(path), data_only=True)
            parts = []
            for name in wb.sheetnames:
                ws = wb[name]
                rows = ["\t".join(str(c.value or "") for c in row)
                        for row in ws.iter_rows()]
                parts.append(f"=== {name} ===\n" + "\n".join(rows))
            return "\n\n".join(parts) or "(empty workbook)"
        except ImportError:
            return "Error: Excel support requires  pip install openpyxl"
        except Exception as e:
            return f"Error reading Excel: {e}"
 
    if ext == ".csv":
        import csv, io
        try:
            raw    = path.read_text(encoding="utf-8", errors="replace")
            reader = csv.reader(io.StringIO(raw))
            return "\n".join("\t".join(row) for row in reader)
        except Exception as e:
            return f"Error reading CSV: {e}"
 
    try:
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return f"Error reading file: {e}"