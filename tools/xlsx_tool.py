"""
tools/excel_tool.py — Shifu's Excel / Spreadsheet Tool
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Read, write, and edit .xlsx / .csv / .tsv files without
ever leaving the agent loop.

Actions
───────
  read        — dump sheet(s) as JSON rows; preview or full data
  read_csv    — read a .csv / .tsv file as JSON rows
  write       — create a brand-new workbook from JSON data
  edit        — update specific cells in an existing workbook
  add_sheet   — append a new sheet to an existing workbook
  delete_sheet— remove a sheet from an existing workbook
  rename_sheet— rename a sheet
  insert_rows — insert blank rows at a position
  delete_rows — delete a range of rows
  insert_cols — insert blank columns at a position
  delete_cols — delete a range of columns
  get_info    — list sheets, dimensions, named ranges
  to_csv      — export one sheet to CSV
  from_csv    — import a CSV as a new sheet inside an xlsx
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

from __future__ import annotations

import csv
import json
import os
from pathlib import Path
from typing import Any

from langchain_core.tools import tool

# ── optional deps ─────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import pandas as pd
    _HAS_PANDAS = True
except ImportError:
    _HAS_PANDAS = False

PLAYGROUND_DIR = Path("Playground")
PLAYGROUND_DIR.mkdir(exist_ok=True)

# ── helpers ───────────────────────────────────────────────────────────────────

def _resolve(path: str) -> Path:
    """Return an absolute Path, trying Playground/ as a fallback."""
    p = Path(path).expanduser()
    if p.is_absolute():
        return p
    if p.exists():
        return p.resolve()
    candidate = PLAYGROUND_DIR / path
    if candidate.exists():
        return candidate.resolve()
    # For write operations the caller wants the Playground location
    return (PLAYGROUND_DIR / path).resolve()


def _require_openpyxl() -> str | None:
    if not _HAS_OPENPYXL:
        return "❌ openpyxl is not installed. Run: pip install openpyxl"
    return None


def _require_pandas() -> str | None:
    if not _HAS_PANDAS:
        return "❌ pandas is not installed. Run: pip install pandas openpyxl"
    return None


def _load_wb(path: Path, data_only: bool = True):
    return load_workbook(str(path), data_only=data_only)


def _sheet(wb, sheet_name: str):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        return wb[sheet_name]
    return wb.active


def _rows_to_json(ws, max_rows: int | None = None) -> list[dict]:
    """Convert a worksheet to a list of dicts (first row = header)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
    data_rows = rows[1:] if len(rows) > 1 else []
    if max_rows is not None:
        data_rows = data_rows[:max_rows]
    result = []
    for row in data_rows:
        result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return result


def _apply_cell_style(cell, style: dict):
    """Apply a style dict to a cell. Keys: bold, italic, color, bg_color, align."""
    font_kwargs: dict[str, Any] = {}
    if style.get("bold"):
        font_kwargs["bold"] = True
    if style.get("italic"):
        font_kwargs["italic"] = True
    if style.get("color"):
        font_kwargs["color"] = style["color"].lstrip("#")
    if font_kwargs:
        cell.font = Font(**font_kwargs)
    if style.get("bg_color"):
        cell.fill = PatternFill("solid", start_color=style["bg_color"].lstrip("#"))
    if style.get("align"):
        cell.alignment = Alignment(horizontal=style["align"])


# ── main tool ─────────────────────────────────────────────────────────────────

@tool
def excel_tool(
    action: str,
    path: str,
    sheet_name: str = "",
    data: str = "[]",
    cell: str = "",
    value: str = "",
    start_row: int = 1,
    num_rows: int = 0,
    start_col: int = 1,
    num_cols: int = 0,
    new_name: str = "",
    preview: int = 20,
    output_path: str = "",
    style_json: str = "{}",
) -> str:
    """Read, write, and edit Excel (.xlsx) and CSV files.

    Args:
        action:       One of: read | read_csv | write | edit | add_sheet |
                      delete_sheet | rename_sheet | insert_rows | delete_rows |
                      insert_cols | delete_cols | get_info | to_csv | from_csv

        path:         Path to the Excel or CSV file.
                      Relative paths are resolved inside Playground/.

        sheet_name:   Target sheet name (default: active/first sheet).

        data:         JSON string used by several actions:
                      • write  — list of dicts (one per row) or
                                 {"sheets": {"Sheet1": [rows...], ...}}
                      • edit   — list of {"cell": "B3", "value": "...", "style": {...}}
                      • from_csv — not used

        cell:         Single cell reference for a quick edit, e.g. "B3".
                      Used together with `value` as a shorthand for edit.

        value:        New value for the cell specified in `cell`.
                      Strings starting with "=" are written as formulas.

        start_row:    1-based row index for insert_rows / delete_rows.
        num_rows:     Number of rows to insert or delete.
        start_col:    1-based column index for insert_cols / delete_cols.
        num_cols:     Number of columns to insert or delete.

        new_name:     New sheet name for rename_sheet, or new sheet name
                      when adding a sheet (add_sheet).

        preview:      Max data rows returned by 'read' (0 = all rows, default 20).

        output_path:  Destination path for 'to_csv' or when 'write' should
                      save to a different location than `path`.

        style_json:   JSON string with style options for the `write` action's
                      header row, e.g. '{"bold": true, "bg_color": "4472C4",
                      "color": "FFFFFF"}'.

    Returns:
        A plain-English status string (or JSON data for read actions).

    Examples
    --------
    # Read first 10 rows from Sheet1
    excel_tool(action="read", path="Playground/data.xlsx",
               sheet_name="Sheet1", preview=10)

    # Write a new file from a list of dicts
    excel_tool(action="write", path="Playground/report.xlsx",
               data='[{"Name":"Alice","Score":95},{"Name":"Bob","Score":88}]')

    # Edit a single cell
    excel_tool(action="edit", path="Playground/report.xlsx",
               cell="B2", value="99")

    # Edit multiple cells at once
    excel_tool(action="edit", path="Playground/report.xlsx",
               data='[{"cell":"A1","value":"Updated"},{"cell":"B1","value":"=SUM(B2:B10)"}]')

    # Add a new sheet
    excel_tool(action="add_sheet", path="Playground/report.xlsx",
               new_name="Summary")

    # Export sheet to CSV
    excel_tool(action="to_csv", path="Playground/report.xlsx",
               output_path="Playground/report.csv")
    """

    # ── route ──────────────────────────────────────────────────────────────────
    action = action.strip().lower()

    if action == "read":
        return _action_read(path, sheet_name, preview)
    elif action == "read_csv":
        return _action_read_csv(path, preview)
    elif action == "write":
        return _action_write(path, data, output_path, style_json)
    elif action == "edit":
        return _action_edit(path, sheet_name, data, cell, value)
    elif action == "add_sheet":
        return _action_add_sheet(path, new_name, data)
    elif action == "delete_sheet":
        return _action_delete_sheet(path, sheet_name)
    elif action == "rename_sheet":
        return _action_rename_sheet(path, sheet_name, new_name)
    elif action == "insert_rows":
        return _action_insert_rows(path, sheet_name, start_row, num_rows)
    elif action == "delete_rows":
        return _action_delete_rows(path, sheet_name, start_row, num_rows)
    elif action == "insert_cols":
        return _action_insert_cols(path, sheet_name, start_col, num_cols)
    elif action == "delete_cols":
        return _action_delete_cols(path, sheet_name, start_col, num_cols)
    elif action == "get_info":
        return _action_get_info(path)
    elif action == "to_csv":
        return _action_to_csv(path, sheet_name, output_path)
    elif action == "from_csv":
        return _action_from_csv(path, sheet_name or new_name, output_path)
    else:
        valid = ("read, read_csv, write, edit, add_sheet, delete_sheet, "
                 "rename_sheet, insert_rows, delete_rows, insert_cols, "
                 "delete_cols, get_info, to_csv, from_csv")
        return f"❌ Unknown action '{action}'. Valid actions: {valid}"


# ── action implementations ────────────────────────────────────────────────────

def _action_read(path: str, sheet_name: str, preview: int) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        wb = _load_wb(p, data_only=True)
        ws = _sheet(wb, sheet_name)
        max_r = None if preview == 0 else preview
        rows = _rows_to_json(ws, max_rows=max_r)
        total = ws.max_row - 1 if ws.max_row and ws.max_row > 1 else 0
        note = f" (showing {len(rows)} of {total} data rows)" if preview and total > preview else ""
        return (
            f"✅ Read sheet '{ws.title}' from {p.name}{note}\n"
            + json.dumps(rows, default=str, indent=2)
        )
    except Exception as exc:
        return f"❌ read failed: {exc}"


def _action_read_csv(path: str, preview: int) -> str:
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        delimiter = "\t" if p.suffix.lower() == ".tsv" else ","
        rows = []
        with open(p, newline="", encoding="utf-8-sig") as fh:
            reader = csv.DictReader(fh, delimiter=delimiter)
            for i, row in enumerate(reader):
                if preview and i >= preview:
                    break
                rows.append(dict(row))
        return f"✅ Read {len(rows)} rows from {p.name}\n" + json.dumps(rows, indent=2)
    except Exception as exc:
        return f"❌ read_csv failed: {exc}"


def _action_write(path: str, data: str, output_path: str, style_json: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        payload = json.loads(data) if data.strip() not in ("", "[]") else []
        style = json.loads(style_json) if style_json.strip() not in ("{}", "") else {}

        dest = _resolve(output_path or path)
        dest.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        # Multi-sheet format: {"sheets": {"Sheet1": [rows...], ...}}
        if isinstance(payload, dict) and "sheets" in payload:
            wb.remove(wb.active)
            for sname, rows in payload["sheets"].items():
                ws = wb.create_sheet(title=sname)
                _write_rows_to_sheet(ws, rows, style)
        else:
            ws = wb.active
            ws.title = "Sheet1"
            _write_rows_to_sheet(ws, payload, style)

        wb.save(str(dest))
        return f"✅ Workbook written: {dest}"
    except Exception as exc:
        return f"❌ write failed: {exc}"


def _write_rows_to_sheet(ws, rows: list[dict], header_style: dict):
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    # Style header row
    if header_style:
        for cell in ws[1]:
            _apply_cell_style(cell, header_style)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    # Auto-size columns
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 60)


def _action_edit(path: str, sheet_name: str, data: str, cell: str, value: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)

        edits: list[dict] = []

        # Shorthand: single cell + value
        if cell.strip():
            edits.append({"cell": cell.strip(), "value": value})

        # Bulk edits from data JSON
        if data.strip() not in ("[]", "", "{}"):
            parsed = json.loads(data)
            if isinstance(parsed, list):
                edits.extend(parsed)
            elif isinstance(parsed, dict):
                edits.append(parsed)

        if not edits:
            return "❌ Provide either cell+value or a data JSON list of {cell, value} objects."

        changed = []
        for item in edits:
            c_ref = str(item.get("cell", "")).strip()
            v = item.get("value", "")
            if not c_ref:
                continue
            # Convert "=..." strings as formulas; everything else as raw value
            ws[c_ref] = v  # openpyxl treats strings starting with "=" as formulas
            changed.append(c_ref)
            if "style" in item and item["style"]:
                _apply_cell_style(ws[c_ref], item["style"])

        wb.save(str(p))
        return f"✅ Updated {len(changed)} cell(s) in '{ws.title}': {', '.join(changed)}"
    except Exception as exc:
        return f"❌ edit failed: {exc}"


def _action_add_sheet(path: str, new_name: str, data: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        wb = _load_wb(p, data_only=False)
        title = new_name.strip() or f"Sheet{len(wb.sheetnames) + 1}"
        if title in wb.sheetnames:
            return f"❌ Sheet '{title}' already exists."
        ws = wb.create_sheet(title=title)
        if data.strip() not in ("[]", "", "{}"):
            rows = json.loads(data)
            if isinstance(rows, list) and rows:
                _write_rows_to_sheet(ws, rows, {})
        wb.save(str(p))
        return f"✅ Added sheet '{title}' to {p.name}"
    except Exception as exc:
        return f"❌ add_sheet failed: {exc}"


def _action_delete_sheet(path: str, sheet_name: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        if not sheet_name:
            return "❌ Provide sheet_name to delete."
        wb = _load_wb(p, data_only=False)
        if sheet_name not in wb.sheetnames:
            return f"❌ Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        del wb[sheet_name]
        wb.save(str(p))
        return f"✅ Deleted sheet '{sheet_name}' from {p.name}"
    except Exception as exc:
        return f"❌ delete_sheet failed: {exc}"


def _action_rename_sheet(path: str, sheet_name: str, new_name: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        if not sheet_name or not new_name:
            return "❌ Provide both sheet_name (current) and new_name."
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)
        ws.title = new_name
        wb.save(str(p))
        return f"✅ Renamed sheet '{sheet_name}' → '{new_name}' in {p.name}"
    except Exception as exc:
        return f"❌ rename_sheet failed: {exc}"


def _action_insert_rows(path: str, sheet_name: str, start_row: int, num_rows: int) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        n = max(num_rows, 1)
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)
        ws.insert_rows(start_row, amount=n)
        wb.save(str(p))
        return f"✅ Inserted {n} row(s) at row {start_row} in '{ws.title}'"
    except Exception as exc:
        return f"❌ insert_rows failed: {exc}"


def _action_delete_rows(path: str, sheet_name: str, start_row: int, num_rows: int) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        n = max(num_rows, 1)
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)
        ws.delete_rows(start_row, amount=n)
        wb.save(str(p))
        return f"✅ Deleted {n} row(s) starting at row {start_row} in '{ws.title}'"
    except Exception as exc:
        return f"❌ delete_rows failed: {exc}"


def _action_insert_cols(path: str, sheet_name: str, start_col: int, num_cols: int) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        n = max(num_cols, 1)
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)
        ws.insert_cols(start_col, amount=n)
        wb.save(str(p))
        col_letter = get_column_letter(start_col)
        return f"✅ Inserted {n} column(s) at column {col_letter} in '{ws.title}'"
    except Exception as exc:
        return f"❌ insert_cols failed: {exc}"


def _action_delete_cols(path: str, sheet_name: str, start_col: int, num_cols: int) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        n = max(num_cols, 1)
        wb = _load_wb(p, data_only=False)
        ws = _sheet(wb, sheet_name)
        ws.delete_cols(start_col, amount=n)
        wb.save(str(p))
        col_letter = get_column_letter(start_col)
        return f"✅ Deleted {n} column(s) starting at column {col_letter} in '{ws.title}'"
    except Exception as exc:
        return f"❌ delete_cols failed: {exc}"


def _action_get_info(path: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        wb = _load_wb(p, data_only=True)
        info: dict[str, Any] = {
            "file": str(p),
            "sheets": {},
        }
        for name in wb.sheetnames:
            ws = wb[name]
            info["sheets"][name] = {
                "rows": ws.max_row,
                "columns": ws.max_column,
                "dimensions": ws.dimensions,
            }
        if wb.defined_names:
            info["named_ranges"] = list(wb.defined_names.keys())
        return "✅ Workbook info:\n" + json.dumps(info, indent=2)
    except Exception as exc:
        return f"❌ get_info failed: {exc}"


def _action_to_csv(path: str, sheet_name: str, output_path: str) -> str:
    err = _require_openpyxl()
    if err:
        return err
    try:
        p = _resolve(path)
        if not p.exists():
            return f"❌ File not found: {p}"
        wb = _load_wb(p, data_only=True)
        ws = _sheet(wb, sheet_name)
        if output_path.strip():
            dest = _resolve(output_path)
        else:
            dest = p.with_suffix(".csv")
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([("" if v is None else v) for v in row])
        return f"✅ Exported sheet '{ws.title}' to CSV: {dest}"
    except Exception as exc:
        return f"❌ to_csv failed: {exc}"


def _action_from_csv(path: str, sheet_name: str, output_path: str) -> str:
    """Import a CSV file as a new sheet inside an xlsx workbook."""
    err = _require_openpyxl()
    if err:
        return err
    try:
        csv_path = _resolve(path)
        if not csv_path.exists():
            return f"❌ CSV file not found: {csv_path}"
        dest_str = output_path.strip()
        if not dest_str:
            return "❌ Provide output_path pointing to the target .xlsx file."
        dest = _resolve(dest_str)
        dest.parent.mkdir(parents=True, exist_ok=True)

        # Load or create the workbook
        if dest.exists():
            wb = _load_wb(dest, data_only=False)
        else:
            wb = Workbook()
            wb.remove(wb.active)

        title = sheet_name.strip() or csv_path.stem
        if title in wb.sheetnames:
            title = f"{title}_imported"
        ws = wb.create_sheet(title=title)

        delimiter = "\t" if csv_path.suffix.lower() == ".tsv" else ","
        with open(csv_path, newline="", encoding="utf-8-sig") as fh:
            for row in csv.reader(fh, delimiter=delimiter):
                ws.append(row)

        wb.save(str(dest))
        return f"✅ Imported '{csv_path.name}' as sheet '{title}' into {dest}"
    except Exception as exc:
        return f"❌ from_csv failed: {exc}"